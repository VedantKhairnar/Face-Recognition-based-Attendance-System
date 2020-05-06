import datetime
from datetime import datetime

import cv2
import face_recognition
import numpy as np
import openpyxl


def addInExcel(d):
    fp = "Attendance.xlsx"
    wb = openpyxl.load_workbook(fp)
    sheet = wb.get_active_sheet()

    max_row = sheet.max_row
    max_column = sheet.max_column

    now = datetime.now()
    # print("Data Acquired :", d)
    sheet.cell(row=1, column=max_column + 1).value = now.strftime("%m/%d/%Y, %H:%M:%S")
    for index in range(1, max_row):
        sheet.cell(row=index + 1, column=max_column + 1).value = d[sheet.cell(row=index + 1, column=2).value]
    wb.save(fp)


video_capture = cv2.VideoCapture(0)

obama_image = face_recognition.load_image_file("images/obama.jpg")
obama_face_encoding = face_recognition.face_encodings(obama_image)[0]
gates_image = face_recognition.load_image_file("images/gates.jpg")
gates_face_encoding = face_recognition.face_encodings(gates_image)[0]

known_face_encodings = [
    obama_face_encoding,
    gates_face_encoding
]
known_face_names = [
    "Barack Obama",
    "Bill Gates"
]
attendance_list = {"Barack Obama": "Absent", "Bill Gates": "Absent", " ": " "}
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

while True:
    ret, frame = video_capture.read()
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = small_frame[:, :, ::-1]

    if process_this_frame:
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
        face_names = []
        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"
            # # If a match was found in known_face_encodings, just use the first one.
            # if True in matches:
            #     first_match_index = matches.index(True)
            #     name = known_face_names[first_match_index]
            # Or instead, use the known face with the smallest distance to the new face
            face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
            best_match_index = np.argmin(face_distances)
            if matches[best_match_index]:
                name = known_face_names[best_match_index]
            if name != "Unknown":
                face_names.append(name)
                attendance_list[name] = "Present"
    process_this_frame = not process_this_frame

    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    cv2.imshow('Video', frame)

    # Hit 'esc' on the keyboard to quit!
    k = cv2.waitKey(30) & 0xff
    if k == 27:
        present = 0
        absent = 0
        print("------------------------------------------")
        print("Attendance as per", datetime.now())
        print("Name  Absent/Present")
        for i in attendance_list:
            print(i, ":", attendance_list[i])

            if attendance_list[i] == "Present":
                present += 1
            elif attendance_list[i] == "Absent":
                absent += 1
        print("Total Present :", present)
        print("Total Absent :", absent)
        print("------------------------------------------")
        attendance_list['Total Present'] = present
        attendance_list['Total Absent'] = absent

        # print(attendance_list)
        addInExcel(attendance_list)
        break

cv2.destroyAllWindows()
video_capture.release()

"""
------------------------------------------
Attendance as per 2020-05-06 09:48:11.096916
Name  Absent/Present
Barack Obama : Present
Bill Gates : Absent
  :  
Total Present : 1
Total Absent : 1
------------------------------------------

"""
